"""
Extract /1/ and /2/ (and 'other') URLs from the sitemaps and write to Excel.
Uses SitemapEN.aspx as the source since all sitemaps share the same hreflang links.
"""

import re
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SITEMAP = os.path.join(os.path.dirname(__file__), "SitemapEN.aspx")
OUTPUT = os.path.join(os.path.dirname(__file__), "pages_list.xlsx")

LANGUAGES = ["en", "de", "fr", "es", "it", "nl"]
LANG_LABELS = {"en": "EN", "de": "DE", "fr": "FR", "es": "ES", "it": "IT", "nl": "NL"}

# Header style
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Type fill colours
FILL_1 = PatternFill("solid", fgColor="D9E1F2")   # light blue  – homepage
FILL_2 = PatternFill("solid", fgColor="E2EFDA")   # light green – info pages
FILL_X = PatternFill("solid", fgColor="FCE4D6")   # light orange – other/unknown


def parse_url_blocks(path):
    with open(path, encoding="utf-8") as fh:
        content = fh.read()

    # Split on <url> blocks
    blocks = re.findall(r"<url>(.*?)</url>", content, re.DOTALL)

    results = []
    for block in blocks:
        loc_match = re.search(r"<loc>(.*?)</loc>", block)
        loc = loc_match.group(1).strip() if loc_match else ""

        # All hreflang alternates
        alternates = {}
        for m in re.finditer(
            r'hreflang="([^"]+)"\s+href="([^"]+)"', block
        ):
            lang = m.group(1).strip().lower()
            url = m.group(2).strip()
            if lang in LANGUAGES:
                alternates[lang] = url

        # Determine type from any URL in the block
        all_urls = list(alternates.values()) + ([loc] if loc else [])
        seg = None
        for u in all_urls:
            m = re.search(r"/(\d+)/", u)
            if m:
                seg = m.group(1)
                break

        if seg in ("1", "2"):
            page_type = seg
        elif seg in ("3", "4"):
            page_type = seg  # shouldn't appear, but mark it
        else:
            page_type = "other"

        results.append({"loc": loc, "type": page_type, "alternates": alternates})

    return results


def slug_from_url(url):
    """Extract the slug (last path segment without .aspx)."""
    m = re.search(r"/([^/]+?)(?:\.aspx)?$", url)
    return m.group(1) if m else url


def build_sheet(ws, lang, pages):
    # Column headers
    headers = ["Type", "URL", "Slug"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for p in pages:
        url = p["alternates"].get(lang, "")
        if not url:
            continue
        t = p["type"]
        slug = slug_from_url(url)
        fill = FILL_1 if t == "1" else (FILL_2 if t == "2" else FILL_X)

        for col, val in enumerate([t, url, slug], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 50


def main():
    blocks = parse_url_blocks(SITEMAP)

    # Keep only /1/, /2/, and "other" (exclude /3/ and /4/)
    filtered = [b for b in blocks if b["type"] not in ("3", "4")]

    # Deduplicate by EN URL (the root homepage block and /1/home.aspx can both resolve to same URL)
    seen = set()
    deduped = []
    for p in filtered:
        key = p["alternates"].get("en") or p["loc"]
        if key not in seen:
            seen.add(key)
            deduped.append(p)
    filtered = deduped

    # Sort: type 1 first, then 2, then other; alpha within each
    def sort_key(p):
        return (p["type"], p["alternates"].get("en", p["loc"]))

    filtered.sort(key=sort_key)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for lang in LANGUAGES:
        ws = wb.create_sheet(title=LANG_LABELS[lang])
        build_sheet(ws, lang, filtered)

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")

    # Summary
    types = defaultdict(int)
    for p in filtered:
        types[p["type"]] += 1
    print(f"  /1/ pages : {types['1']}")
    print(f"  /2/ pages : {types['2']}")
    print(f"  other     : {types['other']}")


if __name__ == "__main__":
    main()
