"""
clean_base.py — Step 1 of the translation pipeline.

Part A: Clear any pre-existing values in 'Translated content'.
Part B: Delete non-spec rows from the base file.

Non-spec categories (deleted):
  json_dict              — Any JSON dict (app config, plugin data, rich text, block refs)
                           Product spec metafields use arrays or primitives, never dicts.
  json_array_en_category — JSON arrays containing English category tag phrases
                           (e.g. ["Candles and candle holders","Organic products"])
  json_array_channel     — JSON arrays whose elements are all sales-channel / audience words
                           (Shop, Webshop, Wholesale, Yoga studio, Yoga, Meditation centre)
  plain_shopify_gid      — Shopify global ID references (gid://shopify/...)
  plain_seo_title        — Contains SEO keywords (groothandel, bestel online, online bestellen,
                           b2b, phoenix import, leverancier) or ` | ` separator pattern
  plain_long_text        — Plain text longer than 100 chars
  plain_language_tag     — Language availability codes like NL/EN/DE/IT/FR/ES
  plain_sku              — Product identifier codes like 01647831L

Usage:
  python translations/clean_base.py
"""

import json
import re
import shutil
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

BASE_FILE = Path("translations/Mani_Bhadra_BV_-_Phoenix_Import_translations_Apr-22-2026.xlsx")
BACKUP_FILE = Path("translations/Mani_Bhadra_BV_-_Phoenix_Import_translations_Apr-22-2026_ORIGINAL.xlsx")

SEO_KEYWORDS = ["groothandel", "bestel online", "online bestellen", "b2b", "phoenix import", "leverancier", " | "]

# English words that appear in category tag metafields but never in Dutch product specs.
EN_CATEGORY_INDICATORS = [
    " and ", " or ", " with ", " for ", " of ", " the ",
    "products", "lifestyle", "holders", "gemstones", "jewelry",
    "organic", "candles", "statues", "stationery", "incense", "singing",
]

# Sales-channel / audience classification values — these are grouping tags, not product specs.
# A JSON array is deleted only when ALL its elements (lowercased) are in this set.
CHANNEL_WORDS = {"shop", "webshop", "wholesale", "yoga studio", "yoga", "meditation centre"}

RE_LANGUAGE_TAG = re.compile(r"^[A-Z]{2}(/[A-Z]{2})+$")
RE_SKU = re.compile(r"^[A-Z]{0,3}\d{5,}[A-Z]{0,2}$")
RE_SHOPIFY_GID = re.compile(r"^gid://shopify/")


def classify(value: str) -> str:
    """Return the delete-category for this value, or '' if the row should be kept."""
    v = value.strip()
    if not v:
        return ""  # blank default content — keep (edge case)

    # Try JSON first
    try:
        parsed = json.loads(v)
        if isinstance(parsed, dict):
            return "json_dict"
        if isinstance(parsed, list):
            v_lower = v.lower()
            if any(kw in v_lower for kw in EN_CATEGORY_INDICATORS):
                return "json_array_en_category"
            if all(isinstance(el, str) and el.lower() in CHANNEL_WORDS for el in parsed):
                return "json_array_channel"
        return ""
    except (json.JSONDecodeError, ValueError):
        pass

    # Plain text checks
    if RE_SHOPIFY_GID.match(v):
        return "plain_shopify_gid"

    v_lower = v.lower()
    for kw in SEO_KEYWORDS:
        if kw in v_lower:
            return "plain_seo_title"

    if len(v) > 100:
        return "plain_long_text"

    if RE_LANGUAGE_TAG.match(v):
        return "plain_language_tag"

    if RE_SKU.match(v):
        return "plain_sku"

    return ""  # keep


def main():
    if not BASE_FILE.exists() and not BACKUP_FILE.exists():
        print(f"ERROR: Neither base file nor backup found.")
        return

    # --- Backup ---
    if BACKUP_FILE.exists():
        print(f"Backup already exists: {BACKUP_FILE}")
    else:
        shutil.copy2(BASE_FILE, BACKUP_FILE)
        print(f"Backup created: {BACKUP_FILE}")

    # Always read from backup so Part A + Part B are both applied cleanly
    source = BACKUP_FILE if BACKUP_FILE.exists() else BASE_FILE
    print(f"\nLoading: {source}")

    # Use read_only for fast streaming read
    wb_src = load_workbook(source, read_only=True)
    ws_src = wb_src.active

    # --- Identify column indices (1-based) ---
    header_row = next(ws_src.iter_rows(min_row=1, max_row=1, values_only=True))
    header = {cell: idx + 1 for idx, cell in enumerate(header_row) if cell is not None}
    required = {"Default content", "Translated content"}
    missing = required - set(header)
    if missing:
        print(f"ERROR: Missing columns: {missing}")
        wb_src.close()
        return

    col_default = header["Default content"]
    col_translated = header["Translated content"]

    # --- Stream rows: classify, count, collect "keep" rows ---
    total_rows = 0
    cleared = 0
    counts = {}
    kept_rows = []  # list of row tuples (values only) with Translated content blanked

    print("Scanning rows...", flush=True)
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        val = row[col_default - 1]
        value = str(val).strip() if val is not None else ""
        category = classify(value)

        if category:
            counts[category] = counts.get(category, 0) + 1
            continue  # skip (delete)

        # Keep this row — blank 'Translated content' (Part A)
        row_list = list(row)
        # Pad short rows to header width
        num_cols = len(header_row)
        while len(row_list) < num_cols:
            row_list.append(None)
        if row_list[col_translated - 1] not in (None, ""):
            row_list[col_translated - 1] = None
            cleared += 1
        kept_rows.append(tuple(row_list))

    wb_src.close()
    print(f"Rows scanned: {total_rows}")
    print(f"\nPart A — Pre-existing translations cleared: {cleared}")

    deleted_total = sum(counts.values())
    remaining = total_rows - deleted_total

    print(f"\nPart B — Rows deleted by category:")
    for cat, n in sorted(counts.items()):
        print(f"  {cat:<28} {n:>6}")
    print(f"  {'TOTAL deleted':<28} {deleted_total:>6}")
    print(f"  {'Rows remaining':<28} {remaining:>6}")

    # --- Write output using write_only mode (fast streaming write) ---
    print(f"\nWriting cleaned file...", flush=True)
    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet()

    # Write header
    ws_out.append(list(header_row))

    # Write kept rows
    for row in kept_rows:
        ws_out.append(list(row))

    wb_out.save(BASE_FILE)
    print(f"Saved cleaned file: {BASE_FILE}")
    print("\n⏸  Please visually inspect the cleaned file before running translate_specs.py.")


if __name__ == "__main__":
    main()
