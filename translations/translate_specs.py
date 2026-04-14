"""
translate_specs.py — Step 2 of the translation pipeline.

Reads the cleaned base xlsx and produces a translated output file for one language.
Every row in the cleaned base is a product spec — there are no rows to skip.

Row classification:
  json_array          Parses as JSON list          → Translate inner text
  json_numeric        Parses as JSON number         → Copy as-is
  plain_country_code  Regex ^[A-Z]{2}$             → Copy as-is (ISO origin code)
  plain_numeric_dim   All digits/separators        → Copy as-is
  plain_spec          Everything else              → Translate

Modes:
  --repair   Claude is asked to repair encoding corruption before translating
  (default)  Corrupted strings passed as-is; Claude translates best-effort

Usage:
  python translations/translate_specs.py --lang EN --repair   → ENspecs.xlsx
  python translations/translate_specs.py --lang EN            → ENspecs_raw.xlsx
  python translations/translate_specs.py --lang DE --repair   → DEspecs.xlsx
  ...

Uses the local `claude` CLI for API calls — no ANTHROPIC_API_KEY needed.
"""

import argparse
import json
import re
import subprocess
import sys
import time
from pathlib import Path

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

BASE_FILE = Path("translations/Mani_Bhadra_BV_-_Phoenix_Import_translations_Apr-13-2026.xlsx")
OUTPUT_DIR = Path("translations")

LANGUAGE_CONFIG = {
    "EN": {"locale": "en", "name": "English"},
    "DE": {"locale": "de", "name": "German (Deutsch)"},
    "FR": {"locale": "fr", "name": "French (Français)"},
    "IT": {"locale": "it", "name": "Italian (Italiano)"},
    "ES": {"locale": "es", "name": "Spanish (Español)"},
}

BATCH_SIZE = 100
MAX_RETRIES = 3
MODEL = "claude-haiku-4-5-20251001"

RE_COUNTRY_CODE = re.compile(r"^[A-Z]{2}$")
RE_NUMERIC_DIM = re.compile(r"^[\d\s.,\-x×\/+%*]+$")

# ---------------------------------------------------------------------------
# Row classification
# ---------------------------------------------------------------------------

def classify_row(value: str) -> str:
    """Return one of: json_array, json_numeric, plain_country_code,
    plain_numeric_dim, plain_spec."""
    v = value.strip()
    try:
        parsed = json.loads(v)
        if isinstance(parsed, list):
            return "json_array"
        if isinstance(parsed, (int, float)):
            return "json_numeric"
    except (json.JSONDecodeError, ValueError):
        pass

    if RE_COUNTRY_CODE.match(v):
        return "plain_country_code"
    if RE_NUMERIC_DIM.match(v):
        return "plain_numeric_dim"
    return "plain_spec"


def extract_translatable(value: str, category: str) -> str | None:
    """Return the string to send to the API, or None if copy-as-is."""
    if category in ("json_numeric", "plain_country_code", "plain_numeric_dim"):
        return None
    if category == "json_array":
        try:
            items = json.loads(value.strip())
            return "\n".join(str(i) for i in items)
        except Exception:
            return value.strip()
    return value.strip()  # plain_spec


def repack_translation(original: str, category: str, translated: str) -> str:
    """Re-wrap a translated string back into its original JSON format."""
    if category == "json_array":
        try:
            original_items = json.loads(original.strip())
            translated_lines = [l.strip() for l in translated.strip().splitlines()]
            if len(translated_lines) == len(original_items):
                return json.dumps(translated_lines, ensure_ascii=False)
            # Count mismatch — wrap whole translation as single-element list
            return json.dumps([translated.strip()], ensure_ascii=False)
        except Exception:
            return translated
    return translated


# ---------------------------------------------------------------------------
# System prompt
# ---------------------------------------------------------------------------

def build_system_prompt(lang_name: str, repair: bool) -> str:
    repair_instruction = (
        "\n- Some inputs contain encoding corruption shown as \ufffd or \ufffd\ufffd\ufffd "
        "where Dutch special characters (ë, é, ij, ö, ü, etc.) should be. "
        "Infer the intended character from context, repair it, then translate the repaired text."
        if repair else ""
    )
    return (
        f"You are translating product specification values for a Dutch B2B wholesale shop "
        f"selling spiritual and new-age products (Phoenix Import). "
        f"Source language: Dutch. Target language: {lang_name}.\n\n"
        f"Rules:\n"
        f"- Translate material names, packaging types, attribute labels, country names, "
        f"colour names, product descriptors, scent names, and size labels.\n"
        f"- Keep proper nouns (brand names, product names) only if they have no standard translation.\n"
        f"- For multi-line inputs representing JSON array elements: translate each line "
        f"independently and return the same number of lines in the same order.\n"
        f"- Keep measurement units (cm, ml, g, kg, mm, st) unchanged.\n"
        f"- Translate country names to their {lang_name} equivalents "
        f"(e.g. India → Indien in German, Inde in French).\n"
        f"- Do NOT add explanations, notes, or extra text.{repair_instruction}\n"
        f"- Return ONLY the translated values, one per line, in the same numbered order as input.\n"
        f"- Format: each line starts with its number and a period, e.g. \"1. translated value\"."
    )


# ---------------------------------------------------------------------------
# Translation via claude CLI
# ---------------------------------------------------------------------------

def translate_batch(
    strings: list[str],
    system_prompt: str,
) -> dict[str, str]:
    """Translate a list of unique strings using the claude CLI.
    Returns {original: translated}."""
    numbered = "\n".join(f"{i+1}. {s}" for i, s in enumerate(strings))

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            result = subprocess.run(
                [
                    "claude", "--print",
                    "--model", MODEL,
                    "--system-prompt", system_prompt,
                    "--tools", "",
                    "--no-session-persistence",
                ],
                input=numbered,
                capture_output=True,
                text=True,
                timeout=120,
            )
            if result.returncode != 0:
                raise RuntimeError(f"claude CLI error: {result.stderr.strip()}")
            raw = result.stdout.strip()
            return parse_numbered_response(strings, raw)
        except Exception as e:
            if attempt == MAX_RETRIES:
                raise
            wait = 2 ** attempt
            print(f"    Error (attempt {attempt}/{MAX_RETRIES}): {e} — retrying in {wait}s")
            time.sleep(wait)

    return {}


def parse_numbered_response(originals: list[str], raw: str) -> dict[str, str]:
    """Parse numbered API response back into {original: translated}."""
    result = {}
    lines = raw.splitlines()
    parsed = []
    for line in lines:
        m = re.match(r"^\s*(\d+)[.)]\s+(.*)", line)
        if m:
            parsed.append((int(m.group(1)), m.group(2).strip()))

    for idx, translation in parsed:
        if 1 <= idx <= len(originals):
            result[originals[idx - 1]] = translation

    # Fill any missing with original (safety net)
    for s in originals:
        if s not in result:
            result[s] = s

    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Translate Shopify spec metafields")
    parser.add_argument("--lang", required=True, choices=LANGUAGE_CONFIG.keys(),
                        help="Target language code (EN/DE/FR/IT/ES)")
    parser.add_argument("--repair", action="store_true",
                        help="Ask Claude to repair encoding corruption before translating")
    args = parser.parse_args()

    lang = args.lang
    repair = args.repair
    cfg = LANGUAGE_CONFIG[lang]
    locale = cfg["locale"]
    lang_name = cfg["name"]

    suffix = "specs" if repair else "specs_raw"
    out_file = OUTPUT_DIR / f"{lang}{suffix}.xlsx"

    print(f"\nTranslating → {lang} ({'repair' if repair else 'raw'})")
    print(f"Output: {out_file}")

    # --- Load base file ---
    print(f"Loading base file: {BASE_FILE}")
    wb = load_workbook(BASE_FILE)
    ws = wb.active

    header = {cell.value: cell.column for cell in ws[1]}
    col_default = header["Default content"]
    col_translated = header["Translated content"]
    col_locale = header["Locale"]

    total_rows = ws.max_row - 1
    print(f"Rows to process: {total_rows}")

    # --- Classify all rows ---
    rows_data = []  # list of (row_num, value, category)
    for row in ws.iter_rows(min_row=2, values_only=False):
        val = row[col_default - 1].value
        value = str(val).strip() if val is not None else ""
        cat = classify_row(value)
        rows_data.append((row[0].row, value, cat))

    # --- Collect unique translatable strings ---
    unique_strings: dict[str, tuple] = {}  # translatable_text → (original_value, category)
    for _, value, cat in rows_data:
        text = extract_translatable(value, cat)
        if text is not None and text not in unique_strings:
            unique_strings[text] = (value, cat)

    translatable = list(unique_strings.keys())
    print(f"Unique translatable strings: {len(translatable)}")

    # --- Translate in batches ---
    system_prompt = build_system_prompt(lang_name, repair)
    translations: dict[str, str] = {}  # translatable_text → translated_text

    total_batches = (len(translatable) + BATCH_SIZE - 1) // BATCH_SIZE
    for i in range(0, len(translatable), BATCH_SIZE):
        batch = translatable[i : i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        print(f"  Batch {batch_num}/{total_batches} ({len(batch)} strings)...", end=" ", flush=True)
        batch_result = translate_batch(batch, system_prompt)
        translations.update(batch_result)
        print("done")

    print(f"Translation complete. {len(translations)} strings translated.")

    # --- Write output file ---
    copy_count = 0
    translate_count = 0

    for row_num, value, cat in rows_data:
        text = extract_translatable(value, cat)
        ws_row = ws[row_num]
        translated_cell = ws_row[col_translated - 1]
        locale_cell = ws_row[col_locale - 1]

        locale_cell.value = locale

        if text is None:
            # Copy as-is
            translated_cell.value = value
            copy_count += 1
        else:
            translated_text = translations.get(text, value)
            translated_cell.value = repack_translation(value, cat, translated_text)
            translate_count += 1

    wb.save(out_file)
    print(f"\nRows copied as-is : {copy_count}")
    print(f"Rows translated    : {translate_count}")
    print(f"Saved: {out_file}")


if __name__ == "__main__":
    main()
